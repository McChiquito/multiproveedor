from __future__ import annotations

import os, csv
from decimal import Decimal

from django.utils import timezone
from openpyxl import load_workbook
import unicodedata
import re
from catalogo.models import Supplier, Product, SupplierProduct, ImportJob
from .matchers import find_existing_product

def _norm(s: str) -> str:
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # quita acentos
    s = s.replace("/", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def _pick(row: dict, colname: str):
    """Obtiene row[colname] de forma tolerante a acentos, espacios, mayúsculas y recortes."""
    if not colname:
        return ""
    norm_target = _norm(colname)
    # mapa normalizado -> original
    normkeys = { _norm(k): k for k in row.keys() }
    # 1) match exacto normalizado
    if norm_target in normkeys:
        return row.get(normkeys[norm_target], "")
    # 2) por prefijo (soporta encabezados recortados)
    for nk, orig in normkeys.items():
        if nk.startswith(norm_target) or norm_target.startswith(nk):
            return row.get(orig, "")
    return ""

def _rows_from_xlsx(path: str):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, [r if r is not None else "" for r in row]))

def _rows_from_csv(path: str):
    # utf-8-sig para tolerar BOM que agrega Excel al CSV
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield {k: (v or "") for k, v in row.items()}

def _iter_rows(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return _rows_from_xlsx(path)
    if ext == ".csv":
        return _rows_from_csv(path)
    raise ValueError("Formato no soportado: use .xlsx o .csv")

def import_for_supplier(supplier: Supplier, path: str, mapping: dict | None = None) -> ImportJob:
    mapping = mapping or supplier.config or DEFAULT_MAPPINGS.get(supplier.slug, {})
    job = ImportJob.objects.create(supplier=supplier, filename=path)

    created_links = updated_links = created_products = 0
    notes = []  # para registrar filas saltadas u observaciones

    for row in _iter_rows(path):
        sku     = str(_pick(row, mapping.get("sku", ""))).strip()
        mpn     = str(_pick(row, mapping.get("mpn", ""))).strip()
        gtin    = str(_pick(row, mapping.get("gtin", ""))).strip()
        name    = str(_pick(row, mapping.get("name", ""))).strip()
        brand   = str(_pick(row, mapping.get("brand", ""))).strip()
        socket  = str(_pick(row, mapping.get("socket", ""))).strip()
        price_raw = _pick(row, mapping.get("price", "")) or 0
        currency  = str(_pick(row, mapping.get("currency", "")) or "MXN").strip() or "MXN"
                # Normalización precio/stock
        try:
            price = Decimal(str(price_raw).replace(",", "")).quantize(Decimal("0.01"))
        except Exception:
            price = Decimal("0.00")
        try:
            stock = int(float(str(stock_raw)))
        except Exception:
            stock = 0

        # ------------------------------------------------------------------
        # 1) Regla principal: si existe el vínculo por (supplier, supplier_sku),
        #    ACTUALIZAR SIN CREAR NADA.
        # ------------------------------------------------------------------
        if sku:
            sp = SupplierProduct.objects.filter(
                supplier=supplier, supplier_sku=sku
            ).select_related("product").first()

            if sp:
                changed = False
                # campos "volátiles" o informativos del vínculo
                if mpn and sp.mpn != mpn:
                    sp.mpn = mpn; changed = True
                if gtin and sp.gtin != gtin:
                    sp.gtin = gtin; changed = True
                if name and sp.name_in_feed != name:
                    sp.name_in_feed = name; changed = True
                if sp.price != price:
                    sp.price = price; changed = True
                if sp.currency != currency:
                    sp.currency = currency; changed = True
                if sp.stock != stock:
                    sp.stock = stock; changed = True

                if changed:
                    sp.save()
                    updated_links += 1

                job.processed_rows += 1
                continue  # ya actualizamos esta fila; seguir con la siguiente

        # ------------------------------------------------------------------
        # 2) Si no existía por SKU, intentamos empatar PRODUCTO por GTIN/MPN/Nombre
        # ------------------------------------------------------------------
        product = find_existing_product(gtin=gtin, mpn=mpn, name=name, socket=socket)

        # Si no hay datos mínimos para crear nada, saltamos
        if not product and not any([sku, mpn, gtin, name]):
            notes.append("Fila sin claves (SKU/MPN/GTIN/Nombre), saltada.")
            job.processed_rows += 1
            continue

        # Crear producto canónico si no existe
        if not product:
            product = Product.objects.create(
                name=name or sku or mpn or gtin or "Producto sin nombre",
                brand=brand,
                mpn=mpn,
                gtin=gtin,
                socket=socket.upper() if socket else "",
                short_description=name[:280] if name else "",
            )
            created_products += 1

        # ------------------------------------------------------------------
        # 3) Crear el vínculo SupplierProduct (o actualizar si ya hay por MPN/GTIN)
        # ------------------------------------------------------------------
        sp, created = SupplierProduct.objects.get_or_create(
            supplier=supplier,
            product=product,
            defaults={
                "supplier_sku": sku,
                "mpn": mpn,
                "gtin": gtin,
                "name_in_feed": name,
                "price": price,
                "currency": currency,
                "stock": stock,
                "availability": "",
            },
        )

        if not created:
            changed = False
            # Si el vínculo existe (quizá por MPN/GTIN), aseguramos que quede el SKU y datos frescos
            if sku and sp.supplier_sku != sku:
                sp.supplier_sku = sku; changed = True
            if mpn and sp.mpn != mpn:
                sp.mpn = mpn; changed = True
            if gtin and sp.gtin != gtin:
                sp.gtin = gtin; changed = True
            if name and sp.name_in_feed != name:
                sp.name_in_feed = name; changed = True
            if sp.price != price:
                sp.price = price; changed = True
            if sp.currency != currency:
                sp.currency = currency; changed = True
            if sp.stock != stock:
                sp.stock = stock; changed = True
            if changed:
                sp.save()
                updated_links += 1
        else:
            created_links += 1

        job.processed_rows += 1

    # Cerrar job
    job.created_links = created_links
    job.updated_links = updated_links
    job.created_products = created_products
    if notes:
        job.notes = "\n".join(notes)
    job.finished_at = timezone.now()
    job.save()
    return job
